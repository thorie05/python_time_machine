from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pandas as pd


ROUND_TO = 6


def read_xlsx(path):
    """Reads a .xlsx luminescence data file."""

    dfs = pd.read_excel(path, sheet_name=None, engine="openpyxl")

    result = {}
    for sheet_name, df in dfs.items():
        columns = df.columns.tolist()[:3]
        arrays = []

        for col in columns:
            arrays.append(df[col].to_numpy(dtype=float))

        while len(arrays) < 3:
            arrays.append(None)

        result[sheet_name] = tuple(arrays)

    return result


def write_xlsx(path, model_function, fit_quality, x_data, y_data, y_err_std,
    bounds, initial_guess, known_params, known_params_err_std,
    free_params_priors, best_fit, confidence_interval, std, rmse,
    bleaching_depth=None):
    """Writes various fit details into a .xlsx file."""

    p = Path(path)
    # ensure directory exists
    p.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active

    current_row = 1

    def _write_title_and_header(title, dct):
        """Helper function that writes a title and a header row."""

        nonlocal current_row

        # bold title cell
        ws.cell(row=current_row, column=1, value=title).font = Font(bold=True)
        current_row += 1

        # parameter names as headers in the row below
        ws.cell(row=current_row, column=1, value="param_name")
        for i, key in enumerate(dct.keys(), start=2):
            ws.cell(row=current_row, column=i, value=key)
        current_row += 1

    def _write_tuple_dict(title, dct, first_name, second_name):
        nonlocal current_row
        """Writes the contents of a dict containing tuples with two elements."""

        # create title and header
        _write_title_and_header(title, dct)

        # create labels for the first and second tuple element
        ws.cell(row=current_row, column=1, value=first_name)
        ws.cell(row=current_row + 1, column=1, value=second_name)

        # fill the two rows with the dict content
        ws.cell(row=current_row, column=1, value=first_name)
        for i, (lower, upper) in enumerate(dct.values(), start=2):
            ws.cell(row=current_row, column=i, value=round(float(lower),
                ROUND_TO))
            ws.cell(row=current_row + 1, column=i, value=round(float(upper),
                ROUND_TO))

        current_row += 3

    def _write_float_and_array_dict(title, dct, value_label="value"):
        """Writes the contents of a dict containing either floats or arrays."""

        nonlocal current_row

        # create title and header
        _write_title_and_header(title, dct)

        # store the maximum row that was written to, to set the new current row
        max_row = current_row 
        for i, value in enumerate(dct.values(), start=2):

            # if value is array write all elements of that array in a column
            if isinstance(value, np.ndarray):
                for j, elem in enumerate(value, start=current_row):
                    ws.cell(row=j, column=i, value=round(float(elem), ROUND_TO))
                    max_row = max(max_row, j)
            # if value is a float just write it
            else:
                ws.cell(row=current_row, column=i, value=round(float(value),
                    ROUND_TO))
                max_row = max(max_row, current_row)

        # if at least one value is an array
        if max_row > current_row:
            ws.cell(row=current_row, column=1, value=value_label+"(s)")
        # if all values are floats
        else:
            ws.cell(row=current_row, column=1, value=value_label)

        current_row = max_row + 2

    ws["A1"] = "MCMC Fit Results"
    ws["A1"].font = Font(bold=True)

    ws["A3"] = "model_function"
    ws["A3"].font = Font(bold=True)
    ws["A4"] = model_function.__name__

    ws["A6"] = "fit_quality"
    ws["A6"].font = Font(bold=True)
    quality_params \
        = ["draws", "tune", "target_accept", "n_bootstrap", "num_restarts"]
    for i, label in enumerate(quality_params, start=1):
        ws.cell(row=7, column=i, value=label)
        val = getattr(fit_quality, label)
        ws.cell(row=8, column=i, value=val)

    ws["A10"] = "Input Data"
    ws["A10"].font = Font(bold=True)
    ws["A11"] = "x_data"
    ws["B11"] = "y_data"
    ws["C11"] = "y_err_std"
    for i, (x, y, y_err) in enumerate(zip(x_data, y_data, y_err_std)):
        ws.cell(row=12 + i, column=1, value=round(float(x), ROUND_TO))
        ws.cell(row=12 + i, column=2, value=round(float(y), ROUND_TO))
        ws.cell(row=12 + i, column=3, value=round(float(y_err), ROUND_TO))

    current_row = 13 + len(x_data) + 1

    _write_tuple_dict("bounds", bounds, "lower", "upper")
    _write_float_and_array_dict("initial_guess", initial_guess)
    _write_float_and_array_dict("known_params", known_params)
    _write_float_and_array_dict("known_params_err_std", known_params_err_std)
    _write_tuple_dict("free_params_priors", free_params_priors, "value", "std")
    _write_float_and_array_dict("best_fit", best_fit, value_label="median")
    _write_tuple_dict("confidence_interval", confidence_interval, "lower",
        "upper")
    _write_float_and_array_dict("std", std)

    # manually add rmse section
    rmse_title_row = current_row
    rmse_value_row = rmse_title_row + 1
    ws.cell(row=rmse_title_row, column=1, value="RMSE").font = Font(bold=True)
    ws.cell(row=rmse_value_row, column=1, value="value")
    ws.cell(row=rmse_value_row, column=2, value=round(float(rmse), ROUND_TO))
    current_row = rmse_value_row + 2

    # manually add bleaching depth if given
    if bleaching_depth is not None:
        bleaching_depth_title_row = current_row
        bleaching_depth_value_row = bleaching_depth_title_row + 1
        ws.cell(row=bleaching_depth_title_row, column=1,
            value="Bleaching-front depth").font = Font(bold=True)
        ws.cell(row=bleaching_depth_value_row, column=1, value="value")
        ws.cell(row=bleaching_depth_value_row, column=2,
            value=round(float(bleaching_depth), ROUND_TO))
        current_row = bleaching_depth_value_row + 2

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(p.as_posix())
